import sys, os, matplotlib, parsefile
import numpy as np
from openpyxl.utils.dataframe import dataframe_to_rows
from profiel import profieldictionary

from openpyxl import load_workbook
from datetime import datetime
from interface import Ui_Interface
from save_pop_up_ui import Ui_Popup
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QTableWidgetItem, QDialog

matplotlib.use("Qt5Agg")


class Save_pop_up(QDialog):
    # just start settings
    def __init__(self, *args):
        QDialog.__init__(self)
        self.ui = Ui_Popup()
        self.ui.setupUi(self)
        self.setWindowTitle("Saving_file_settings")

        self.ui.pushButton.clicked.connect(self.reject)
        self.ui.Save.clicked.connect(self.accept)
        self.file_name = ''
        self.file_location = os.path.join(os.path.expanduser("~"), "Downloads")
        self.ui.FileLocation.setText(self.file_location)
        self.ui.SearchFile.clicked.connect(self.openFileDialog)
        self.ui.FileName.setText(f"Overzichtstabel {datetime.now().strftime('%Y-%m-%d %H-%M-%S')}")

    def openFileDialog(self):
      file_dialog = QFileDialog(self)
      file_dialog.setWindowTitle("Open File")
      file_dialog.setFileMode(QFileDialog.FileMode.Directory)
      file_dialog.setViewMode(QFileDialog.ViewMode.List)

      if file_dialog.exec():
         selected_files = file_dialog.selectedFiles()
         self.ui.FileLocation.setText(selected_files[0])
         self.file_location = selected_files[0]

    # returns the file_location and FileName.
    def get_values(self):
        return self.file_location, self.ui.FileName.text()

class Interface(QMainWindow):
    def __init__(self, *args):
        super().__init__()
        self.ui = Ui_Interface()
        self.ui.setupUi(self)
        self.setWindowTitle("Stempelframe")

        self.ui.ExtractWalls.setDisabled(True)
        self.ui.ProcessFile.setDisabled(True)
        self.ui.ExtractStempels.setDisabled(True)
        self.ui.ExportStempels.setDisabled(True)
        self.ui.GordingUnityCheck.setDisabled(True)
        self.ui.SearchFile.clicked.connect(self.Searchfile)
        self.ui.ProcessFile.clicked.connect(self.Processfile)
        self.ui.TableLists.currentIndexChanged.connect(self.Selectionchanged)
        self.ui.ExtractStempels.clicked.connect(self.ExtractStempels)
        self.ui.ExportStempels.clicked.connect(self.Savedata)
        self.ui.ExtractWalls.clicked.connect(self.Extractwalls)
        self.ui.GordingUnityCheck.clicked.connect(self.GordingUnityCheck)

        self.dataframes = {}
        self.tables = []

    def GordingUnityCheck(self):
        self.walls = self.Extractwalls()
        for wall in self.walls:
            for rod in wall:
                rod_dict = self.fill_in_rod_properties(rod)

    def fill_in_rod_properties(self, rod):
        property_dict = {}
        property_dict, (aantal, profiel) = self.propertiesMaterialen(property_dict, rod)
        property_dict = self.propertiesProfiel(property_dict, profiel, aantal)
        property_dict = self.propertiesBelastingen(property_dict)
        property_dict = self.propertiesClassificatie(property_dict)
        property_dict = self.toetsingDoorsnede(property_dict)
        print(property_dict)

    def toetsingDoorsnede(self, property_dict):
        property_dict = self.toetsingNormaalkracht(property_dict)
        property_dict = self.toetsingBuigendMoment(property_dict)
        property_dict = self.toetsingDwarskracht(property_dict)
        property_dict = self.toetsingBuigingEnDwarskracht(property_dict)
        property_dict = self.toetsingBuigingEnNormaalkracht(property_dict)
        return self.toetsingBuigingEnDwarskrachtEnNormaalkracht(property_dict)

    def toetsingBuigingEnDwarskrachtEnNormaalkracht(self, property_dict):
        if property_dict["doorsnede klasse druk en buiging"]() < 3:
            if property_dict["u.c dwarskracht"] <= 0.5:
                property_dict["rho2"] = lambda: 0
            else:
                property_dict["rho2"] = lambda: ((((2*property_dict["UGT Dwarskracht"])/property_dict["aantal"])/property_dict["Rek. dwarskracht vloeien"]())-1)**2
            property_dict["NVz,Rd"] = lambda: ((property_dict["A"]*property_dict["fy"])-(property_dict["rho2"]()*property_dict["Av"]*property_dict["fy"]))/(1*1000)
            if ((property_dict["A"]-(2*property_dict["b"]*property_dict["tf"]))/property_dict["A"])<0.5:
                property_dict["a1"] = ((property_dict["A"]-(2*property_dict["b"]*property_dict["tf"]))/property_dict["A"])
            else:
                property_dict["a1"] = 0.5
            property_dict["a2"] = lambda: property_dict["a1"]*(1-property_dict["rho2"]())
            property_dict["MN,y,Rd"] = lambda: ((property_dict["Wy-y"]-((property_dict["rho2"]*property_dict["Av"]^2)/(4*property_dict["tw"])))*property_dict["fy"]/1)/10**6
        else:
            pass
        return property_dict

    def toetsingBuigingEnNormaalkracht(self, property_dict):
        if property_dict["doorsnede klasse druk en buiging"]() < 3:
            property_dict["NEd/hw*tw*fy/2*gm0"] = lambda: (property_dict["Rek. normaalkracht"]()*1000)/(property_dict["Hw"]*property_dict["tw"]*property_dict["fy"])/2
            property_dict["a"] = (property_dict["A"]-(2*property_dict["b"]*property_dict["tf"]))/property_dict["A"]
            if property_dict["u.c normaalkracht"] <= 0.25 and property_dict["NEd/hw*tw*fy/2*gm0"] <= 1:
                property_dict["u.c buiging en normaalkracht"] = lambda: property_dict["Rek. weerstand tegen buigend moment"]
            else:
                property_dict["u.c buiging en normaalkracht"] = lambda: property_dict["Wy-y"]*property_dict["fy"]*((1-property_dict["u.c normaalkracht"]())/(1-(0.5*property_dict["a"])))/1000000
        else:
            property_dict["u.c buiging en normaalkracht"] = lambda: (property_dict["Rek. normaalkracht"]()*1000)/(property_dict["A"]*property_dict["fy"])+(property_dict["Rek. buigend moment"]()*1000000/(property_dict["wel"]*property_dict["fy"]))
        return property_dict

    def toetsingBuigingEnDwarskracht(self, property_dict):
        if property_dict["u.c dwarskracht"]() <= 0.5:
            property_dict["rho"] = 0
        else:
            property_dict["rho"] = lambda: property_dict["UGT Dwarskracht"]*1.1/property_dict["aantal"]
        property_dict["My,V,Rd"] = lambda: ((property_dict["Wy-y"]-((property_dict["rho"]*(property_dict["Aw"]**2))/(4*property_dict["tw"])))*property_dict["fy"]/1)/1000000
        property_dict["u.c buiging en normaalkracht"] = lambda: property_dict["My,V,Rd"]()/property_dict["Rek. buigendmoment"]()

    def toetsingDwarskracht(self, property_dict):
        property_dict["Rek. dwarskracht"] = lambda: property_dict["UGT Dwarskracht"]*1.1/property_dict["aantal"]
        if property_dict["doorsnede klasse druk"] <= 2:
            property_dict["Rek. dwarskracht vloeien"] = property_dict["Av"]*property_dict["S"]/1/1000
        else:
            property_dict["Rek. dwarskracht vloeien"] = property_dict["Aw"]*property_dict["S"]/1/1000
        property_dict["u.c dwarskracht"] = lambda: property_dict["Rek. dwarskracht"]()/ property_dict["Rek. dwarskracht vloeien"]()

    def toetsingBuigendMoment(self, property_dict):
        property_dict["Rek. buigendmoment"] = lambda: property_dict["UGT Moment"]*1.1/property_dict["aantal"]
        if property_dict["doorsnede klasse buiging"] <= 2:
            property_dict["Wy-y"] = property_dict["wpl"]
            property_dict["Rek. weerstand tegen buigend moment"] = lambda: (property_dict["fy"]*property_dict["wpl"])/1/1000000
        else:
            property_dict["Wy-y"] = property_dict["wel"]
            property_dict["Rek. weerstand tegen buigend moment"] = lambda: (property_dict["fy"]*property_dict["wel"])/1/1000000
        property_dict["u.c buigend moment"] = lambda: property_dict["Rek. buigendmoment"]()/ property_dict["Rek. weerstand tegen buigend moment"]()

    def toetsingNormaalkracht(self, property_dict):
        property_dict["Rek. normaalkracht"] = lambda: property_dict["UGT Normaalkracht"]*1.1/property_dict["aantal"]
        property_dict["Rek. weerstand tegen trek"] = lambda: (property_dict["fy"]*property_dict["A"])/1/1000
        property_dict["u.c normaalkracht"] = lambda: property_dict["Rek. normaalkracht"]()/ property_dict["Rek. weerstand tegen trek"]()

    def propertiesBelastingen(self, property_dict):
        property_dict["UGT Normaalkracht"] = 0
        property_dict["UGT Dwarskracht"] = 0
        property_dict["UGT Moment"] = 0

        return property_dict

    def propertiesClassificatie(self, property_dict):
        property_dict["lijf"] = (property_dict["h"]-2*property_dict["tf"]-2*property_dict["r"])/property_dict["tw"]/property_dict["epsilon"]
        property_dict["flens"] = (property_dict["b"]/2 -(property_dict["tw"]/2)-property_dict["r"])/property_dict["tf"]/property_dict["epsilon"]
        if property_dict["h/b"] > 1.2 and property_dict["tf"] <= 40:
            property_dict["knikkromme y-y"] = "a"
            property_dict["knikkromme z-z"] = "b"
        elif property_dict["h/b"] > 1.2 and property_dict["tf"] > 40 and property_dict["tf"] <= 100:
            property_dict["knikkromme y-y"] = "b"
            property_dict["knikkromme z-z"] = "c"
        elif property_dict["h/b"] <= 1.2 and property_dict["tf"] <= 100:
            property_dict["knikkromme y-y"] = "b"
            property_dict["knikkromme z-z"] = "c"
        elif property_dict["h/b"] <= 1.2 and property_dict["tf"] > 100:
            property_dict["knikkromme y-y"] = "d"
            property_dict["knikkromme z-z"] = "d"
        else:
            raise Exception("either h/b is not correct or tf is not correctly check them again")

        if property_dict["lijf"] < 72:
            property_dict["doorsnede klasse buiging"] = 1
        elif property_dict["lijf"] < 83:
            property_dict["doorsnede klasse buiging"] = 2
        elif property_dict["lijf"] < 124:
            property_dict["doorsnede klasse buiging"] = 3
        else:
            property_dict["doorsnede klasse buiging"] = 4

        if property_dict["lijf"] < 33:
            property_dict["doorsnede klasse druk"] = 1
        elif 33 < property_dict["lijf"] < 38:
            property_dict["doorsnede klasse druk"] = 2
        elif 38 < property_dict["lijf"] < 42:
            property_dict["doorsnede klasse druk"] = 3
        elif property_dict["lijf"] > 42:
            property_dict["doorsnede klasse druk"] = 4

        property_dict["temp1"] = lambda: (
        property_dict["lijf"] * ((13 * property_dict["imperfectie"]()) - 1)
        if property_dict["imperfectie"]() > 0.5
        else property_dict["lijf"] * property_dict["imperfectie"]()
        )

        property_dict["temp2"] = lambda: (
            396 if property_dict["imperfectie"]() > 0.5 else 36
        )

        property_dict["temp3"] = lambda: (
            456 if property_dict["imperfectie"]() > 0.5 else 41.5
        )

        property_dict["doorsnede klasse druk en buiging"] = lambda: (
            1 if property_dict["temp1"]() < property_dict["temp2"]() and property_dict["temp1"]() < property_dict["temp3"]()
            else 2 if (property_dict["temp1"]() > 51 and property_dict["temp1"]() < property_dict["temp3"]()) or 
                    (property_dict["temp1"]() < 51 and property_dict["temp1"]() > property_dict["temp3"]())
            else 3 if property_dict["temp1"]() > property_dict["temp2"]() and property_dict["temp1"]() > property_dict["temp3"]() else None)
        return property_dict

    def propertiesProfiel(self, property_dict, profiel, aantal):
        property_dict["Type Gording"] = profiel
        property_dict["aantal"] = aantal
        profielpropertie = profieldictionary[profiel]
        property_dict["h"] = profielpropertie[0]
        property_dict["b"] = profielpropertie[1]
        property_dict["tw"] = profielpropertie[2]
        property_dict["tf"] = profielpropertie[3]
        property_dict["wel"] = profielpropertie[4]
        property_dict["wpl"] = profielpropertie[5]
        property_dict["A"] = profielpropertie[8]
        property_dict["r"] = profielpropertie[9]
        property_dict["ly"] = profielpropertie[10]
        property_dict["G"] = profielpropertie[12]
        property_dict["Hw"] = property_dict["b"]-2*property_dict["tf"]
        property_dict["Av"] = property_dict["A"]-(2*property_dict["b"]*property_dict["tf"])+((property_dict["tw"]+2*property_dict["r"])*property_dict["tf"])
        property_dict["Af"] = property_dict["b"] * property_dict["tf"]
        property_dict["Aw"] = property_dict["tw"] * property_dict["Hw"]
        property_dict["h/b"] = property_dict["h"] / property_dict["b"]
        return property_dict


    def propertiesMaterialen(self, property_dict, rod):
        STAVEN = self.Findtable("STAVEN")
        Nr, Profiel = STAVEN[2][rod-1][3].split(":")
        PROFIEL = self.Findtable("PROFIELEN [mm]")
        property_dict["S"] = int(PROFIEL[2][int(Nr)-1][2].split(":")[1][1:])
        property_dict["fy"] = property_dict["S"]
        match property_dict["fy"]:
            case 235:
                property_dict["fu"] = 360
            case 275:
                property_dict["fu"] = 430
            case 355:
                property_dict["fu"] = 510
            case 440:
                property_dict["fu"] = 550
        property_dict["E"] = 210000
        property_dict["epsilon"] = 235/property_dict["fy"]
        property_dict["lambda"] = float(np.pi*np.sqrt(property_dict["E"]/property_dict["fy"]))
        return property_dict, Profiel.split("*")

    def Extractwalls(self):
        walls = []
        coordinates = self.Getcoordinatelist()
        staafgordingen = self.Getgordingen()

        # Build a mapping from knoop to sticks
        knoop_to_sticks = {}
        for stick in staafgordingen:
            _, p1, p2 = stick
            knoop_to_sticks.setdefault(p1, []).append(stick)
            knoop_to_sticks.setdefault(p2, []).append(stick)

        visited = set()
        current_wall = []

        def get_other_point(stick, known_point):
            return stick[1] if stick[2] == known_point else stick[2]

        for stick in staafgordingen:
            stick_id, p1, p2 = stick
            if stick_id in visited:
                continue

            current_wall = [stick_id]
            visited.add(stick_id)
            current_point = p2
            prev_point = p1

            while True:
                connected_sticks = [s for s in knoop_to_sticks.get(current_point, []) if s[0] not in visited]
                found_next = False
                last_wall = 0

                for next_stick in connected_sticks:
                    next_id, np1, np2 = next_stick
                    next_point = get_other_point(next_stick, current_point)

                    angle = self.angle_at_point(
                        coordinates[current_point - 1],
                        coordinates[prev_point - 1],
                        coordinates[next_point - 1]
                    )
                    if angle > 150:
                        current_wall.append(next_id)
                        visited.add(next_id)
                        prev_point, current_point = current_point, next_point
                        found_next = True
                        break  # Continue chaining
                    else:
                        last_wall = next_id

                if not found_next:
                    if last_wall != 0:
                        current_wall.append(last_wall)
                    else:
                        for rod in staafgordingen:
                            if rod[1] == np2:
                                current_wall.append(rod[0])
                    break

            walls.append(current_wall)
        return walls

    def Getgordingen(self):
        gordingen = []
        table = self.Findtable("STAVEN")
        for item in table[2]:
            if "H" in item[3]:
                gordingen.append(item[:3])
        return gordingen

    def Savedata(self):
        dialog = Save_pop_up()
        if dialog.exec_():
            file_path, file_name = dialog.get_values()
            if getattr(sys, 'frozen', False):
                # Running as a bundled exe
                base_path = sys._MEIPASS
            else:
                # Running in normal Python
                base_path = os.path.dirname(__file__)

            original_path = os.path.join(base_path, 'templates', 'template.xlsx')
            workbook = load_workbook(original_path)

            new_sheet = workbook['Stempels']
            table = self.Findtable("Rapport data")
            dataframe = parsefile.generate_dataframe(table)

            # Write DataFrame to Excel
            for r in dataframe_to_rows(dataframe, index=False, header=True):
                for i, item in enumerate(r):
                    if isinstance(item, str):
                        try:
                            r[i] = abs(float(item.replace(",", ".")))
                        except ValueError:
                            pass
                new_sheet.append(r)
            for row in new_sheet.iter_rows(min_row=2):
                for cell in row:
                    print(cell.value)
                    if isinstance(cell.value, float):
                        cell.number_format = '#.##0,00'
            workbook.save(os.path.join(file_path, f"{file_name}.xlsx"))

    def Findtable(self, tablename):
        for table in self.tables:
            if table[0] == tablename:
                return table
        raise Exception(f"Could find the table {tablename}")

    def Searchfile(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Select File")
        if file_path:
            self.ui.SelectedFile.setText(file_path)
            self.ui.ProcessFile.setEnabled(True)

    def Processfile(self):
        self.parseddata, self.tables = parsefile.parsefile(self.ui.SelectedFile.text())
        self.ui.TableLists.clear()
        self.ui.ExtractStempels.setEnabled(True)
        self.ui.ExtractWalls.setEnabled(True)
        self.ui.GordingUnityCheck.setEnabled(True)
        for table in self.tables:
            self.ui.TableLists.addItem(table[0])
        self.Generategraph()

    def Generategraph(self):
        self.ui.Graph.canvas.axes.clear()
        coordinatelist = self.Getcoordinatelist()
        plotted_labels = set()  # Track which labels have been added

        table = self.Findtable("STAVEN")
        for item in table[2]:
            x_vals = [coordinatelist[item[1]-1][0], coordinatelist[item[2]-1][0]]
            y_vals = [coordinatelist[item[1]-1][1], coordinatelist[item[2]-1][1]]

            if "H" not in item[3]:
                label = "stempel" if "stempel" not in plotted_labels else None
                self.ui.Graph.canvas.axes.plot(x_vals, y_vals, 'g', markersize=1, label=label)
                plotted_labels.add("stempel")
            else:
                label = "gording" if "gording" not in plotted_labels else None
                self.ui.Graph.canvas.axes.plot(x_vals, y_vals, 'b', markersize=1, label=label)
                plotted_labels.add("gording")


        for coordinate in coordinatelist:
            label = "knopen" if "knopen" not in plotted_labels else None
            self.ui.Graph.canvas.axes.plot(coordinate[0], coordinate[1], 'ro', markersize=1, label=label)
            plotted_labels.add("knopen")

        self.ui.Graph.canvas.axes.set_aspect('equal', adjustable='box')
        self.ui.Graph.canvas.axes.legend(loc="upper left")
        self.ui.Graph.canvas.draw()

    def Getcoordinatelist(self):
        coordinatelist = []
        table = self.Findtable("KNOPEN")
        for item in table[2]:
            coordinatelist.append([item[1], item[2]])
        return coordinatelist

    def Selectionchanged(self, index):
        self.ui.TableData.clearContents()
        self.ui.TableData.setRowCount(0)

        for table in self.tables:
            if table[0] == self.ui.TableLists.itemText(index):
                headers = table[1]
                data = table[2]

                self.ui.TableData.setColumnCount(len(headers))
                self.ui.TableData.setHorizontalHeaderLabels([str(h) for h in headers])
                self.ui.TableData.setRowCount(len(data))

                for row_idx, datarow in enumerate(data):
                    for col_idx, item in enumerate(datarow):
                        cell = QTableWidgetItem(str(item))
                        self.ui.TableData.setItem(row_idx, col_idx, cell)

    def ExtractStempels(self):
        self.Generateknopenset()
        self.getprof_len()
        self.getquality()
        self.getforce()
        self.calculateangles()

        name = "Rapport data"
        existing_items = [self.ui.TableLists.itemText(i) for i in range(self.ui.TableLists.count())]
        if name not in existing_items:
            self.ui.TableLists.addItem(name)
            self.tables.append([name, self.knopentable[:1][0], self.knopentable[1:]])
        self.ui.TableLists.setCurrentText(name)
        self.ui.ExportStempels.setEnabled(True)



    def calculateangles(self):
        table = self.Findtable("STAVEN")
        for knoop in self.knopentable[1:]:
            for item in table[2]:
                if item[0] == knoop[0]:
                    startknoop = item[1]
                    eindknoop = item[2]
                    angle1 = self.Findshortestangle(startknoop, eindknoop)
                    angle2 = self.Findshortestangle(eindknoop, startknoop)
                    knoop[5] = angle1
                    knoop[6] = angle2

    def Findshortestangle(self, startknoop, eindknoop):
        linkknopen = [startknoop]
        table = self.Findtable("STAVEN")
        for item in table[2]:
            if "H" in item[3]:
                if item[1] == startknoop:
                    linkknopen.append(item[2])
                elif item[2] == startknoop:
                    linkknopen.append(item[1])
        return self.Getcoordinates(linkknopen, eindknoop)

    def Getcoordinates(self, linkknopen, eindknoop):
        table = self.Findtable("KNOPEN")
        for item in table[2]:
            for num, knoop in enumerate(linkknopen):
                if item[0] == knoop:
                    linkknopen[num] = [item[1], item[2]]
                    continue
                if item[0] == eindknoop:
                    eindknoop = [item[1], item[2]]
        angle_list = []
        for knoop in linkknopen[1:]:
            angle_list.append(self.angle_at_point(linkknopen[0], knoop, eindknoop))

        return float(round(self.get_smallest_angle(angle_list), 1))

    def get_smallest_angle(self, angle_list):
        while len(angle_list) > 2:
            angle_list.remove(max(angle_list))
        if len(angle_list) == 2:
            angle = 0
            if angle_list[0] + angle_list[1] > 90:
                angle = 180 - max(angle_list)
            else:
                angle = min(angle_list)
            if angle < 30:
                angle = 90 - angle
            return angle
        else:
            exit(0)

    def angle_at_point(self, B, A, C):
        BA = [A[0] - B[0], A[1] - B[1]]
        BC = [C[0] - B[0], C[1] - B[1]]
        return self.angle_between_vectors(BA, BC)

    def angle_between_vectors(self, v1, v2):
        v1 = np.array(v1)
        v2 = np.array(v2)
        dot_product = np.dot(v1, v2)
        norm_product = np.linalg.norm(v1) * np.linalg.norm(v2)
        angle_rad = np.arccos(dot_product / norm_product)
        angle_deg = np.degrees(angle_rad)
        return angle_deg

    def getforce(self):
        for knoop in self.knopentable:
            for item in self.knopenset:
                if knoop[0] == item[0]:
                    if item[2] == 0:
                        knoop[7] = item[1]
                    else:
                        knoop[8] = item[1]
                    continue

    def getquality(self):
        table = self.Findtable("PROFIELEN [mm]")
        for knoop in self.knopentable:
            for item in table[2][1:]:
                if knoop[1] in item[1] and knoop[2] in item[1]:
                    knoop[3] = item[2].split(':')[1]
                    continue

    def getprof_len(self):
        self.knopenlist = []
        self.knopentable = [['Nr', chr(216),  'w', "st. kwal.", "Lengte", "Hoek 1", "Hoek 2", "UGT: Sterkte", "BGT: Vervorming"]]
        for item in self.knopenset:
            if item[0] not in self.knopenlist:
                self.knopenlist.append(item[0])
        table = self.Findtable("STAVEN")
        for staaf in table[2]:
            if staaf[0] in self.knopenlist:
                profiel = staaf[3].split('B')[1].split('/')
                lengte = float(staaf[6])
                self.knopentable.append([staaf[0], profiel[0], profiel[1], 0, lengte, 0, 0, 0, 0])

    def Generateknopenset(self):
        self.knopenset = set()
        Krachttable = ["STAAFKRACHTEN  B.C:1 Sterkte", "STAAFKRACHTEN  B.C:2 Vervorming"]
        for table in self.tables:
            if table[0] in Krachttable:
                for item in table[2]:
                    if not item[2] and not item[4] and not item[5] and item [3]:
                        if "Sterkte" in table[0]:
                            knoop = (item[0], item[3], 0)
                        else:
                            knoop = (item[0], item[3], 1)
                        self.knopenset.add(knoop)



if __name__ == "__main__":
    app = QApplication(sys.argv)
    form = Interface()
    form.show()
    sys.exit(app.exec_())